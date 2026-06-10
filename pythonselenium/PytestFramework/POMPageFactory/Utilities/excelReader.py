import openpyxl

def get_data(path,sheetName):
    
    final_list=[]
    Workbook = openpyxl.load_workbook(path)
    sheet = Workbook[sheetName]
    no_of_rows = sheet.max_row
    no_of_columns = sheet.max_column
    
    
    for r in range(2,no_of_rows+1):
        row_list = []
        for c in range(1,no_of_columns+1):
            row_list.append(sheet.cell(r,c).value)
        final_list.append(row_list)
    
    return final_list